import os

import openpyxl


class SourcesLoader:

    EXPECTED_HEADERS = [
        "Active", "Region", "Jurisdiction", "Filename",
        "Validity", "Source URL", "Referer", "Notes",
    ]

    @classmethod
    def load(cls, sources_xlsx_path):

        if not os.path.exists(sources_xlsx_path):
            raise FileNotFoundError(
                f"sources file not found at {sources_xlsx_path}. "
                "Run tools/build_sources_workbook.py to generate it, "
                "or create it manually with columns: "
                + ", ".join(cls.EXPECTED_HEADERS)
            )

        wb = openpyxl.load_workbook(sources_xlsx_path, data_only=True)

        ws = wb.active

        headers = [
            (ws.cell(row=1, column=c).value or "").strip()
            for c in range(1, ws.max_column + 1)
        ]

        idx = {h: i for i, h in enumerate(headers)}

        for required in cls.EXPECTED_HEADERS:

            if required not in idx:
                raise ValueError(
                    f"sources file is missing required column '{required}'"
                )

        india_pdfs = {}
        india_validity = {}
        india_html = {}

        country_pdfs = {}
        country_validity = {}
        country_html = {}

        for r in range(2, ws.max_row + 1):

            row = [ws.cell(row=r, column=c + 1).value for c in range(len(headers))]

            active = str(row[idx["Active"]] or "").strip().lower()

            if active not in ("yes", "y", "true", "1"):
                continue

            region = (row[idx["Region"]] or "").strip()
            jurisdiction = (row[idx["Jurisdiction"]] or "").strip()
            filename = (row[idx["Filename"]] or "").strip()
            validity = (row[idx["Validity"]] or "").strip()
            url = (row[idx["Source URL"]] or "").strip()
            referer = (row[idx["Referer"]] or "").strip() or None

            if not jurisdiction or not url:
                continue

            fmt = cls._url_format(url)

            if region == "India":

                if jurisdiction == "India-State wise":
                    continue  # Just URL aggregation, not a state name

                if validity and jurisdiction not in india_validity:
                    india_validity[jurisdiction] = validity

                if fmt == "PDF" and filename:

                    india_pdfs.setdefault(jurisdiction, []).append(
                        (filename, url, referer)
                    )

                elif fmt == "HTML":

                    india_html.setdefault(jurisdiction, []).append(url)

            else:

                if validity and jurisdiction not in country_validity:
                    country_validity[jurisdiction] = validity

                if fmt == "PDF" and filename:

                    country_pdfs.setdefault(jurisdiction, []).append(
                        (filename, url, referer)
                    )

                elif fmt == "HTML":

                    country_html.setdefault(jurisdiction, []).append(url)

        return {
            "india_pdfs": india_pdfs,
            "india_validity": india_validity,
            "india_html": india_html,
            "country_pdfs": country_pdfs,
            "country_validity": country_validity,
            "country_html": country_html,
        }

    @staticmethod
    def _url_format(url):

        if not url or not url.startswith("http"):
            return ""

        return "PDF" if url.lower().split("?")[0].endswith(".pdf") else "HTML"
