import glob
import os
import re

import openpyxl


class PreviousWorkbookReader:

    BULLET_SPLIT = "\n\n• "

    WHITESPACE_RE = re.compile(r"\s+")

    @classmethod
    def find_latest(cls, output_dir, prefix, exclude_path=None):

        pattern = os.path.join(output_dir, f"{prefix}*.xlsx")

        candidates = [
            p for p in glob.glob(pattern)
            if not os.path.basename(p).startswith("~$")
        ]

        if exclude_path:

            candidates = [
                p for p in candidates
                if os.path.abspath(p) != os.path.abspath(exclude_path)
            ]

        if not candidates:
            return None

        candidates.sort(key=os.path.getmtime, reverse=True)

        return candidates[0]

    @classmethod
    def load(cls, path):

        if not path or not os.path.exists(path):
            return {}

        try:

            wb = openpyxl.load_workbook(path, data_only=True)

        except Exception:

            return {}

        result = {}

        for sn in wb.sheetnames:

            ws = wb[sn]

            cells = {}

            for r in range(1, ws.max_row + 1):

                for c in range(1, ws.max_column + 1):

                    v = ws.cell(r, c).value

                    if isinstance(v, str) and v.strip():
                        cells[(r, c)] = v

            result[sn] = cells

        return result

    @classmethod
    def parse_bullets(cls, text):

        if not text:
            return []

        text = text.strip()

        if text.startswith("• "):
            text = text[2:]

        items = text.split(cls.BULLET_SPLIT)

        return [it.strip() for it in items if it.strip()]

    @classmethod
    def normalize(cls, text):

        return cls.WHITESPACE_RE.sub(" ", text).strip().lower()
