import re

import openpyxl


class ExpectedSourcesReader:

    URL_RE = re.compile(r"https?://[^\s)<>\"']+")

    @classmethod
    def read(cls, expected_xlsx_path):

        wb = openpyxl.load_workbook(expected_xlsx_path, data_only=False)

        result = {}

        for sn in wb.sheetnames:

            ws = wb[sn]

            urls = []

            seen = set()

            for row in ws.iter_rows():

                for cell in row:

                    if cell.hyperlink and cell.hyperlink.target:

                        url = cell.hyperlink.target

                        if url not in seen:
                            seen.add(url)
                            urls.append(url)

                    val = cell.value

                    if val and isinstance(val, str):

                        for m in cls.URL_RE.findall(val):

                            if m not in seen:
                                seen.add(m)
                                urls.append(m)

            if urls:
                result[sn] = urls

        return result
