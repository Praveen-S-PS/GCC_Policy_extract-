from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side


class SourcesExcelExporter:

    HEADERS = [
        "Region",
        "Country / State",
        "Source URL",
        "Format",
        "Status",
        "Local Path",
        "Notes",
    ]

    HEADER_FILL = PatternFill("solid", fgColor="1F3864")
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    GROUP_FILL = PatternFill("solid", fgColor="D9E1F2")
    BORDER = Border(
        left=Side(style="thin", color="BFBFBF"),
        right=Side(style="thin", color="BFBFBF"),
        top=Side(style="thin", color="BFBFBF"),
        bottom=Side(style="thin", color="BFBFBF"),
    )
    WRAP = Alignment(wrap_text=True, vertical="top", horizontal="left")

    @classmethod
    def export(cls, rows, output_path):

        wb = Workbook()

        ws = wb.active

        ws.title = "Sources"

        for c, h in enumerate(cls.HEADERS, start=1):

            cell = ws.cell(row=1, column=c, value=h)
            cell.font = cls.HEADER_FONT
            cell.fill = cls.HEADER_FILL
            cell.alignment = Alignment(
                wrap_text=True, vertical="center", horizontal="center"
            )
            cell.border = cls.BORDER

        for r_idx, row in enumerate(rows, start=2):

            for c_idx, value in enumerate(row, start=1):

                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.alignment = cls.WRAP
                cell.border = cls.BORDER

                if c_idx == 3 and isinstance(value, str) and value.startswith("http"):

                    cell.hyperlink = value
                    cell.font = Font(color="0563C1", underline="single")

        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 75
        ws.column_dimensions["D"].width = 10
        ws.column_dimensions["E"].width = 14
        ws.column_dimensions["F"].width = 40
        ws.column_dimensions["G"].width = 45

        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 28

        for r in range(2, len(rows) + 2):

            ws.row_dimensions[r].height = 38

        wb.save(output_path)
