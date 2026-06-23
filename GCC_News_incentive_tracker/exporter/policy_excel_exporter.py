from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont

from feeds.previous_workbook_reader import PreviousWorkbookReader


class PolicyExcelExporter:

    INCENTIVE_TYPES = [
        "Financial Incentives",
        "Infrastructure",
        "Talent",
        "Regulatory, Ease-of-Doing-Business & Non-financial Support",
        "Innovation / R&D / Patent",
        "Others",
    ]

    SCOPES = [
        "Across the state (Common)",
        "Specific to Tier 1 cities",
        "Specific to Tier 2 & 3 cities",
    ]

    HEADER_FILL = PatternFill(
        "solid", fgColor="1F3864"
    )

    HEADER_FONT = Font(bold=True, color="FFFFFF")

    TYPE_FILL = PatternFill(
        "solid", fgColor="D9E1F2"
    )

    BORDER = Border(
        left=Side(style="thin", color="BFBFBF"),
        right=Side(style="thin", color="BFBFBF"),
        top=Side(style="thin", color="BFBFBF"),
        bottom=Side(style="thin", color="BFBFBF"),
    )

    WRAP = Alignment(
        wrap_text=True, vertical="top", horizontal="left"
    )

    NEW_FONT = InlineFont(rFont="Calibri", sz=11, color="C00000")
    OLD_FONT = InlineFont(rFont="Calibri", sz=11, color="000000")

    @classmethod
    def export(cls, state_data, validity_by_state, output_path):

        wb = Workbook()

        wb.remove(wb.active)

        cls._write_definition_sheet(wb)

        cls._write_india_sheet(wb, state_data, validity_by_state)

        wb.save(output_path)

    @classmethod
    def export_global(
        cls,
        state_data,
        validity_by_state,
        country_data,
        validity_by_country,
        placeholder_data,
        placeholder_validity,
        output_path,
        baseline=None,
    ):

        has_baseline = bool(baseline)

        baseline = baseline or {}

        new_items = []

        wb = Workbook()

        wb.remove(wb.active)

        cls._write_definition_sheet(wb)

        cls._write_india_sheet(
            wb, state_data, validity_by_state,
            baseline=baseline.get("India-State wise", {}),
            has_baseline=has_baseline,
            new_items=new_items,
        )

        for country in country_data:

            cls._write_country_sheet(
                wb,
                country,
                country_data[country],
                validity_by_country.get(country, ""),
                baseline=baseline.get(country[:31], {}),
                has_baseline=has_baseline,
                new_items=new_items,
            )

        for country in placeholder_data:

            cls._write_country_sheet(
                wb,
                country,
                placeholder_data[country],
                placeholder_validity.get(country, ""),
                baseline=baseline.get(country[:31], {}),
                has_baseline=has_baseline,
                new_items=new_items,
            )

        cls._write_summary_sheet(wb, new_items, has_baseline)

        # Put Summary right after Definition
        wb.move_sheet("Summary", offset=-(len(wb.sheetnames) - 2))

        wb.save(output_path)

    @classmethod
    def _write_definition_sheet(cls, wb):

        ws = wb.create_sheet("Definition & Methodology")

        rows = [
            ("Definition:",),
            (
                "Incentive Type: The classification of government support offered to Global Capability Centres (GCCs) based on the nature of assistance provided, including the following broad categories:\n"
                "1. Financial Incentives\n2. Infrastructure\n3. Talent\n"
                "4. Regulatory, Ease-of-Doing-Business & Non-financial Support\n"
                "5. Innovation / R&D / Patent\n6. Others",
            ),
            (
                "Applicable in Which Part of the State refers to the geographical scope within a state where a specific GCC incentive is valid. Usage Classification:\n"
                "• Across the state (Common): uniformly applicable across all eligible locations.\n"
                "• Specific to Tier 1 cities: applicable only to designated Tier 1 cities.\n"
                "• Specific to Tier 2 & 3 cities: applicable only to designated Tier 2/3 cities.",
            ),
            (
                "Policy Validity Period: the duration during which a government policy remains in force.",
            ),
            (
                "Caveat/Conditions: requirements, limitations, or stipulations attached to availing a benefit.",
            ),
            ("",),
            ("Methodology:",),
            ("This document presents GCC incentives extracted from official state policy PDFs.",),
            ("Country Coverage (v1): India only, 8 priority states.",),
        ]

        for r, row in enumerate(rows, start=1):

            ws.cell(row=r, column=2, value=row[0]).alignment = cls.WRAP

        ws.column_dimensions["A"].width = 2
        ws.column_dimensions["B"].width = 120

        for r in range(1, len(rows) + 1):

            ws.row_dimensions[r].height = 60

    @classmethod
    def _write_india_sheet(cls, wb, state_data, validity_by_state, baseline=None, has_baseline=False, new_items=None):

        baseline = baseline or {}

        ws = wb.create_sheet("India-State wise")

        states = list(state_data.keys())

        # Header row
        headers = ["Incentive Type", "Applicable in which part of the state"] + states

        for c, h in enumerate(headers, start=1):

            cell = ws.cell(row=1, column=c, value=h)
            cell.font = cls.HEADER_FONT
            cell.fill = cls.HEADER_FILL
            cell.alignment = Alignment(
                wrap_text=True, vertical="center", horizontal="center"
            )
            cell.border = cls.BORDER

        # Validity row
        ws.cell(row=2, column=1, value="Policy Validity Period ->").font = Font(bold=True)
        ws.cell(row=2, column=1).alignment = cls.WRAP
        ws.cell(row=2, column=1).border = cls.BORDER

        ws.cell(row=2, column=2, value="").border = cls.BORDER

        for col_idx, state in enumerate(states, start=3):

            cell = ws.cell(
                row=2, column=col_idx,
                value=validity_by_state.get(state, "")
            )
            cell.alignment = cls.WRAP
            cell.border = cls.BORDER

        # Incentive rows
        row_cursor = 3

        for incentive_type in cls.INCENTIVE_TYPES:

            type_start_row = row_cursor

            for scope in cls.SCOPES:

                type_cell = ws.cell(row=row_cursor, column=1)

                if scope == cls.SCOPES[0]:

                    type_cell.value = incentive_type
                    type_cell.font = Font(bold=True)
                    type_cell.fill = cls.TYPE_FILL

                type_cell.alignment = cls.WRAP
                type_cell.border = cls.BORDER

                scope_cell = ws.cell(row=row_cursor, column=2, value=scope)
                scope_cell.alignment = cls.WRAP
                scope_cell.border = cls.BORDER

                for col_idx, state in enumerate(states, start=3):

                    items = (
                        state_data
                        .get(state, {})
                        .get(incentive_type, {})
                        .get(scope, [])
                    )

                    cls._write_diff_cell(
                        ws, row_cursor, col_idx, items,
                        baseline.get((row_cursor, col_idx), ""),
                        has_baseline=has_baseline,
                        new_items=new_items,
                        context={
                            "region": "India",
                            "sheet": "India-State wise",
                            "jurisdiction": state,
                            "incentive_type": incentive_type,
                            "scope": scope,
                        },
                    )

                row_cursor += 1

            if type_start_row + 1 <= row_cursor - 1:

                ws.merge_cells(
                    start_row=type_start_row,
                    start_column=1,
                    end_row=row_cursor - 1,
                    end_column=1,
                )

        # Column widths
        ws.column_dimensions["A"].width = 24
        ws.column_dimensions["B"].width = 28

        for col_idx in range(3, 3 + len(states)):

            col_letter = ws.cell(row=1, column=col_idx).column_letter

            ws.column_dimensions[col_letter].width = 55

        ws.row_dimensions[1].height = 30
        ws.row_dimensions[2].height = 30

        for r in range(3, row_cursor):

            ws.row_dimensions[r].height = 220

    @classmethod
    def _write_country_sheet(cls, wb, country, buckets, validity, baseline=None, has_baseline=False, new_items=None):

        baseline = baseline or {}

        ws = wb.create_sheet(country[:31])

        headers = [
            "Incentive Type",
            "Applicable in which part of the Country",
            country,
        ]

        for c, h in enumerate(headers, start=1):

            cell = ws.cell(row=1, column=c, value=h)
            cell.font = cls.HEADER_FONT
            cell.fill = cls.HEADER_FILL
            cell.alignment = Alignment(
                wrap_text=True, vertical="center", horizontal="center"
            )
            cell.border = cls.BORDER

        ws.cell(row=2, column=1, value="Policy Validity Period ->").font = Font(bold=True)
        ws.cell(row=2, column=1).alignment = cls.WRAP
        ws.cell(row=2, column=1).border = cls.BORDER
        ws.cell(row=2, column=2).border = cls.BORDER

        cell = ws.cell(row=2, column=3, value=validity)
        cell.alignment = cls.WRAP
        cell.border = cls.BORDER

        country_scopes = [
            "Across the Country (Common)",
            "Specific to Tier 1 cities",
            "Specific to Tier 2 & 3 cities",
        ]

        row_cursor = 3

        for incentive_type in cls.INCENTIVE_TYPES:

            type_start_row = row_cursor

            for idx, scope in enumerate(country_scopes):

                type_cell = ws.cell(row=row_cursor, column=1)

                if idx == 0:

                    type_cell.value = incentive_type
                    type_cell.font = Font(bold=True)
                    type_cell.fill = cls.TYPE_FILL

                type_cell.alignment = cls.WRAP
                type_cell.border = cls.BORDER

                scope_cell = ws.cell(row=row_cursor, column=2, value=scope)
                scope_cell.alignment = cls.WRAP
                scope_cell.border = cls.BORDER

                bucketer_scope = cls.SCOPES[idx]

                items = buckets.get(incentive_type, {}).get(bucketer_scope, [])

                cls._write_diff_cell(
                    ws, row_cursor, 3, items,
                    baseline.get((row_cursor, 3), ""),
                    has_baseline=has_baseline,
                    new_items=new_items,
                    context={
                        "region": "Global",
                        "sheet": country,
                        "jurisdiction": country,
                        "incentive_type": incentive_type,
                        "scope": scope,
                    },
                )

                row_cursor += 1

            if type_start_row + 1 <= row_cursor - 1:

                ws.merge_cells(
                    start_row=type_start_row,
                    start_column=1,
                    end_row=row_cursor - 1,
                    end_column=1,
                )

        ws.column_dimensions["A"].width = 24
        ws.column_dimensions["B"].width = 28
        ws.column_dimensions["C"].width = 80

        ws.row_dimensions[1].height = 30
        ws.row_dimensions[2].height = 30

        for r in range(3, row_cursor):

            ws.row_dimensions[r].height = 220

    @classmethod
    def _write_summary_sheet(cls, wb, new_items, has_baseline):

        ws = wb.create_sheet("Summary")

        headers = [
            "Region", "Sheet", "Jurisdiction",
            "Incentive Type", "Scope", "New Content",
        ]

        for c, h in enumerate(headers, start=1):

            cell = ws.cell(row=1, column=c, value=h)
            cell.font = cls.HEADER_FONT
            cell.fill = cls.HEADER_FILL
            cell.alignment = Alignment(
                wrap_text=True, vertical="center", horizontal="center"
            )
            cell.border = cls.BORDER

        if not has_baseline:

            note = (
                "No prior workbook found in output/ folder. "
                "Run the script a second time to start producing change summaries."
            )

            ws.cell(row=2, column=1, value=note).alignment = cls.WRAP

            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)

            ws.row_dimensions[2].height = 40

        elif not new_items:

            note = (
                "No new content versus the previous run. "
                "All cells matched the prior workbook."
            )

            ws.cell(row=2, column=1, value=note).alignment = cls.WRAP

            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)

            ws.row_dimensions[2].height = 30

        else:

            row_fill = PatternFill("solid", fgColor="FCE4E4")
            red_font = Font(color="C00000", bold=False)

            for r_idx, item in enumerate(new_items, start=2):

                values = [
                    item.get("region", ""),
                    item.get("sheet", ""),
                    item.get("jurisdiction", ""),
                    item.get("incentive_type", ""),
                    item.get("scope", ""),
                    item.get("text", ""),
                ]

                for c_idx, v in enumerate(values, start=1):

                    cell = ws.cell(row=r_idx, column=c_idx, value=v)
                    cell.alignment = cls.WRAP
                    cell.border = cls.BORDER
                    cell.fill = row_fill

                    if c_idx == 6:
                        cell.font = red_font

                ws.row_dimensions[r_idx].height = 60

        ws.column_dimensions["A"].width = 9
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 28
        ws.column_dimensions["E"].width = 22
        ws.column_dimensions["F"].width = 90

        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 30

    @classmethod
    def _write_placeholder_country_sheet(cls, wb, country):

        empty_buckets = {
            it: {scope: [] for scope in cls.SCOPES}
            for it in cls.INCENTIVE_TYPES
        }

        cls._write_country_sheet(wb, country, empty_buckets, "")

    @classmethod
    def _write_diff_cell(cls, ws, row, col, items, baseline_text, has_baseline=False, new_items=None, context=None):

        deduped = cls._dedupe(items)

        cell = ws.cell(row=row, column=col)
        cell.alignment = cls.WRAP
        cell.border = cls.BORDER

        if not deduped:
            cell.value = ""
            return

        # No prior workbook at all — render plain, no highlighting
        if not has_baseline:

            cell.value = "\n\n• ".join([""] + deduped).strip()
            return

        previous_items = PreviousWorkbookReader.parse_bullets(baseline_text)

        previous_keys = {
            PreviousWorkbookReader.normalize(p) for p in previous_items
        }

        blocks = []

        for i, item in enumerate(deduped):

            prefix = "• " if i == 0 else "\n\n• "

            is_new = PreviousWorkbookReader.normalize(item) not in previous_keys

            font = cls.NEW_FONT if is_new else cls.OLD_FONT

            blocks.append(TextBlock(font, prefix + item))

            if is_new and new_items is not None and context is not None:

                new_items.append({
                    **context,
                    "text": item,
                })

        cell.value = CellRichText(blocks)

    @staticmethod
    def _dedupe(items):

        if not items:
            return []

        seen = set()

        deduped = []

        for it in items:

            key = it[:60]

            if key in seen:
                continue

            seen.add(key)
            deduped.append(it)

        return deduped

    @staticmethod
    def _format_cell(items):

        if not items:
            return ""

        seen = set()
        deduped = []

        for it in items:

            key = it[:60]

            if key in seen:
                continue

            seen.add(key)
            deduped.append(it)

        return "\n\n• ".join([""] + deduped).strip() if deduped else ""
