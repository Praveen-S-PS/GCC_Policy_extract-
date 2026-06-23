"""One time generator for input/sources.xlsx.

Pulls URLs from the human-curated expected workbook plus our previously
hardcoded configs and writes them into a single editable Excel file with
one row per source URL. After running this once, you can edit
input/sources.xlsx to add, remove, or change sources without touching code.

Re-run this only if you want to rebuild from the expected workbook from scratch.
"""

import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from extractor.expected_sources_reader import ExpectedSourcesReader


EXPECTED_WORKBOOK = (
    "/Users/test/Downloads/Ops_Geo_GCCIncentivesPolicyInsights_05212026_v6.xlsx"
)

DEST = "input/sources.xlsx"


PRECONFIGURED_INDIA = [
    # (state, filename, validity, url, referer)
    ("Gujarat", "gujarat.pdf", "Feb 2025 – Mar 2030",
     "https://api.giftgujarat.in/public/GCC_Policy.pdf", None),
    ("Karnataka", "karnataka.pdf", "2024 – 2029",
     "https://eitbt.karnataka.gov.in/uploads/media_to_upload1732030878.pdf", None),
    ("Madhya Pradesh", "madhya_pradesh.pdf", "Feb 2025 – Feb 2030",
     "https://mpsedc.mp.gov.in/Uploaded%20Document/UpcomingEvents/10122024030121Madhya%20Pradesh%20GCC%20Policy%202024.pdf", None),
    ("Maharashtra", "maharashtra.pdf", "Nov 2025 – Mar 2030",
     "https://industry.maharashtra.gov.in/sites/default/files/2025-09/gcc-policy-paraesa-naota-english.pdf", None),
    ("Rajasthan", "rajasthan.pdf", "2025 – Mar 2029",
     "https://avantiscdnprodstorage.blob.core.windows.net/legalupdatedocs/50160/Rajasthan-Govt.-issued-the-Global-Capability-Center-Policy-2025-dec022025.pdf", None),
    ("Tamil Nadu", "tamil_nadu.pdf", "Apr 2024 – Mar 2027", None, None),
    ("Telangana", "telangana.pdf", "2023 – 2028",
     "https://gccrise.com/wp-content/uploads/2025/01/Nasscom-GCC-Telangana-Playbook-Nov-2024.pdf", None),
    ("Uttar Pradesh", "uttar_pradesh.pdf", "Jun 2025 – Jun 2030",
     "https://invest.up.gov.in/wp-content/uploads/2025/06/GCC-Policy-Eng_050625.pdf", None),
]


PRECONFIGURED_COUNTRY_PDFS = [
    ("Mexico", "Jan 2025 – Sept 2030", "mexico_plan_mexico_decreto.pdf",
     "https://www.estimulosfiscales.hacienda.gob.mx/work/models/estimulosfiscales/Resource/3/decreto.pdf",
     "https://www.estimulosfiscales.hacienda.gob.mx/es/efiscales_mediante_decreto/Plan_Mexico"),
    ("Mexico", "Jan 2025 – Sept 2030", "mexico_lfzee.pdf",
     "https://www.diputados.gob.mx/LeyesBiblio/pdf/LFZEE.pdf", None),
    ("Poland", "2018 – ongoing", "poland_paih_bss_2025.pdf",
     "https://www.paih.gov.pl/wp-content/uploads/2026/03/Delivered-from-Poland-2025.pdf", None),
    ("Hungary", "Ongoing", "hungary_state_aid_decision.pdf",
     "https://ec.europa.eu/competition/state_aid/cases1/202144/SA_63934_A00D127C-0100-CE43-A1A3-1DECB7658033_164_1.pdf", None),
    ("Mauritius", "Ongoing", "mauritius_ict_investment_certificate.pdf",
     "https://edbmauritius.org/wp-content/uploads/2022/12/Investment-CertificateICT.pdf", None),
    ("Morocco", "2022 – ongoing", "morocco_investment_charter_en.pdf",
     "https://www.morocconow.com/wp-content/uploads/2023/07/Charte-ANG-06072023.pdf", None),
    ("Panama", "Ongoing", "panama_ley_412_2004.pdf",
     "https://www.mici.gob.pa/wp-content/uploads/2021/11/03-LEY-PP-L412004.pdf", None),
    ("Romania", "Ongoing", "romania_investor_guide_en.pdf",
     "https://arice.gov.ro/1/wp-content/uploads/2024/06/Caietul-Investitorului-04.06.2024-en.pdf", None),
    ("Romania", "Ongoing", "romania_hg332_2014_state_aid.pdf",
     "https://mfinante.gov.ro/static/10/Mfp/buget/sitebuget/ajutordestat/HGnr332_2014establishingStateaidscheme.pdf", None),
    ("South Africa", "Ongoing", "south_africa_gbs_incentive_guidelines.pdf",
     "https://www.thedtic.gov.za/wp-content/uploads/GBS-Incentive-Guidelines.pdf", None),
    ("South Africa", "Ongoing", "south_africa_gbs_notice.pdf",
     "https://www.thedtic.gov.za/wp-content/uploads/Notice-to-GBS-Sector.pdf", None),
    ("Spain", "Ongoing", "spain_state_aid_decision.pdf",
     "https://ec.europa.eu/competition/state_aid/cases1/202213/SA_100859_409A9C7F-0000-C192-80EB-3A23648EC558_56_1.pdf", None),
    ("Spain", "Ongoing", "spain_tax_relief_rd.pdf",
     "https://www.investinspain.org/content/dam/icex-invest/documentos/publicaciones/doing-business/Tax_relief_for_RD_and_innovation_activities_in_Spain.pdf", None),
    ("UAE", "2026 – ongoing", "uae_rd_tax_credit_decision_2026.pdf",
     "https://mof.gov.ae/wp-content/uploads/2026/03/Ministerial-Decision-No.-24-of-2026-on-the-Implementation-of-Certain-Provisions-of-Cabinet-Decision-No.-215-of-2025-on-Research-Development-Tax-Credit-en.pdf", None),
    ("Zimbabwe", "Ongoing", "zimbabwe_investors_handbook.pdf",
     "https://zimtreasury.co.zw/wp-content/uploads/2025/04/The-Investors-Tax-Incentive-Handbook-text.pdf", None),
]


HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF")
BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)


HEADERS = [
    "Active",
    "Region",
    "Jurisdiction",
    "Filename",
    "Validity",
    "Source URL",
    "Referer",
    "Notes",
]


def url_format(url):

    if not url or not url.startswith("http"):
        return ""

    return "PDF" if url.lower().split("?")[0].endswith(".pdf") else "HTML"


def main():

    rows = []

    for state, filename, validity, url, referer in PRECONFIGURED_INDIA:

        rows.append([
            "yes" if url else "no",
            "India",
            state,
            filename or "",
            validity,
            url or "",
            referer or "",
            "Used as PDF source for the India-State wise sheet"
            if url else "No public source located",
        ])

    for country, validity, filename, url, referer in PRECONFIGURED_COUNTRY_PDFS:

        rows.append([
            "yes",
            "Global",
            country,
            filename,
            validity,
            url,
            referer or "",
            "PDF source for the country sheet",
        ])

    # HTML reference URLs from the expected workbook
    if os.path.exists(EXPECTED_WORKBOOK):

        expected = ExpectedSourcesReader.read(EXPECTED_WORKBOOK)

        existing_urls = {r[5] for r in rows}

        for sheet_name, urls in expected.items():

            if sheet_name == "Definition & Methodology":
                continue

            region = "India" if sheet_name == "India-State wise" else "Global"

            for url in urls:

                if url in existing_urls:
                    continue

                rows.append([
                    "yes" if url_format(url) == "HTML" else "no",
                    region,
                    sheet_name,
                    "",
                    "",
                    url,
                    "",
                    "HTML reference from expected workbook"
                    if url_format(url) == "HTML"
                    else "PDF reference, not yet wired in",
                ])

    wb = Workbook()

    ws = wb.active

    ws.title = "Sources"

    for c, h in enumerate(HEADERS, start=1):

        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(
            wrap_text=True, vertical="center", horizontal="center"
        )
        cell.border = BORDER

    for r_idx, row in enumerate(rows, start=2):

        for c_idx, value in enumerate(row, start=1):

            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = Alignment(
                wrap_text=True, vertical="top", horizontal="left"
            )
            cell.border = BORDER

            if c_idx == 6 and isinstance(value, str) and value.startswith("http"):

                cell.hyperlink = value
                cell.font = Font(color="0563C1", underline="single")

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 9
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 32
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 70
    ws.column_dimensions["G"].width = 30
    ws.column_dimensions["H"].width = 45

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30

    for r in range(2, len(rows) + 2):

        ws.row_dimensions[r].height = 30

    os.makedirs("input", exist_ok=True)

    wb.save(DEST)

    print(f"Wrote {DEST} with {len(rows)} source rows")


if __name__ == "__main__":
    main()
